import sys, pandas as pd, numpy as np
sys.stdout.reconfigure(encoding="utf-8")

print("问题4: 产能扩张分析")
print()

WEEKLY_PROD = 28200
CONSUME = {"A": 0.6, "B": 0.66, "C": 0.72}
PRICE = {"A": 1.2, "B": 1.1, "C": 1.0}
PLAN_WEEKS = 24
TRANS_CAP = 6000
N_TRANS = 8

fname = "附件1 近5年402家供应商的相关数据.xlsx"
order_df = pd.read_excel(fname, sheet_name=0, header=0)
supply_df = pd.read_excel(fname, sheet_name=1, header=0)
trans_df = pd.read_excel("附件2 近5年8家转运商的相关数据.xlsx", header=0)
ranking = pd.read_csv("supplier_ranking.csv")
week_cols = [c for c in supply_df.columns if c.startswith("W")]
sup_vals = supply_df[week_cols].values.astype(float)
ord_vals = order_df[week_cols].values.astype(float)
trans_loss = trans_df[week_cols].values.astype(float)
supply_ids = supply_df["供应商ID"].values
cats = supply_df["材料分类"].values

# 最大产能计算
total_supply = sup_vals.sum(axis=0).mean()
cat_supply = {}
cat_prod = {}
for c in ["A", "B", "C"]:
    m = cats == c
    cat_supply[c] = sup_vals[m].sum(axis=0).mean()
    cat_prod[c] = cat_supply[c] / CONSUME[c]
max_prod = sum(cat_prod.values())

print("当前产能: {} m3产品/周".format(WEEKLY_PROD))
print("最大可持续产能: {:.0f} m3产品/周".format(max_prod))
print("产能差距: {:.0f} m3 ({:.1f}%)".format(max_prod - WEEKLY_PROD, (max_prod - WEEKLY_PROD) / WEEKLY_PROD * 100))
print("原因: 原材料供给不足 (总 {} m3/周)".format(total_supply))
print()

# 使用所有供应商, A > B > C 优先
ranking["avg_supply"] = ranking["供应商ID"].map({sid: sup_vals[i].mean() for i, sid in enumerate(supply_ids)})
ranking["材料分类"] = ranking["供应商ID"].map({sid: cats[i] for i, sid in enumerate(supply_ids)})
ranking["mat_priority"] = ranking["材料分类"].map({"A": 0, "B": 1, "C": 2})
ranking_sorted = ranking.sort_values(["mat_priority", "最终得分"], ascending=[True, False])

# 全选402家
selected_sids = list(ranking_sorted["供应商ID"])

sd = {}
for _, row in ranking.iterrows():
    sid = row["供应商ID"]
    si = np.where(supply_ids == sid)[0][0]
    sd[sid] = {"order": row["avg_supply"], "sup_idx": si, "mat": row["材料分类"]}

# ratio 按材料类别取均值(同题2)
sr = np.where(ord_vals > 0, sup_vals / np.maximum(ord_vals, 1), np.nan)
mr = np.nan_to_num(np.nanmedian(sr, axis=1), nan=1.0)
tl = [np.mean(trans_loss[t][trans_loss[t] > 0]) / 100 if len(trans_loss[t][trans_loss[t] > 0]) > 0 else 0 for t in range(N_TRANS)]
avg_loss = np.mean(tl)

orders_by_cat = {}
for cat in ["A", "B", "C"]:
    sub = ranking[ranking["材料分类"] == cat]
    tot = sub["avg_supply"].sum()
    mids = sub["供应商ID"].values
    mask = np.isin(supply_ids, mids)
    r = mr[mask]
    # ar unused: avg_supply already accounts for delivery behavior
    orders_by_cat[cat] = {"order": tot, "supply": tot}

recv_total = sum(v["supply"] for v in orders_by_cat.values()) * (1 - avg_loss)
act_prod = sum(v["supply"] * (1 - avg_loss) / CONSUME[k] for k, v in orders_by_cat.items())

total_order = sum(sd[s]["order"] for s in selected_sids)
print("订购方案: {}家, {:.0f} m3/周".format(len(selected_sids), total_order))

# 运输分配
tsort = sorted(range(N_TRANS), key=lambda t: tl[t])
sl = [(sid, max(1, int(round(sd[sid]["order"] * mr[sd[sid]["sup_idx"]])))) for sid in selected_sids]
sl.sort(key=lambda x: -x[1])
tu = np.zeros(N_TRANS)
ta = {}
for sid, qty in sl:
    for t in tsort:
        if tu[t] + qty <= TRANS_CAP:
            tu[t] += qty; ta[sid] = t; break
    else:
        t = max(range(N_TRANS), key=lambda x: TRANS_CAP - tu[x])
        tu[t] += qty; ta[sid] = t

print("运输分配:")
for t in range(N_TRANS):
    cnt = sum(1 for v in ta.values() if v == t)
    if tu[t] > 0: print("  T{}: {}家, {} m3".format(t + 1, cnt, int(tu[t])))
print("  合计: {}家".format(len(ta)))

print()
print("效果分析:")
print("  周采购成本: {:.0f}".format(sum(sd[s]["order"] * PRICE[sd[s]["mat"]] for s in selected_sids)))
print("  周接收量: {:.0f} m3".format(recv_total))
print("  可实现产量: {:.0f} m3 (满足率 {:.1f}%)".format(act_prod, act_prod / WEEKLY_PROD * 100))
print("  理论极限产能: {:.0f} m3产品/周 (满足率 {:.1f}%)".format(max_prod, max_prod / WEEKLY_PROD * 100))
print("  产能提升: {:.0f} m3 ({:+.1f}%)".format(max_prod - WEEKLY_PROD, (max_prod - WEEKLY_PROD) / WEEKLY_PROD * 100))
print()

# 填充附件A问题4
import openpyxl
wbA = openpyxl.load_workbook("附件A 订购方案数据结果.xlsx")
wsA = wbA["问题4的订购方案结果"]
for r in range(7, 409):
    for c in range(2, 26): wsA.cell(r, c).value = None
for r in range(7, 409):
    sid = wsA.cell(r, 1).value
    if sid and sid in sd:
        o = max(1, int(round(sd[sid]["order"])))
        for c in range(2, 26): wsA.cell(r, c).value = o
wbA.save("附件A 订购方案数据结果.xlsx")
print("附件A Q4 OK")

# 填充附件B问题4
wbB = openpyxl.load_workbook("附件B 转运方案数据结果.xlsx")
wsB = wbB["问题4的转运方案结果"]
for r in range(7, 409):
    for c in range(2, 194): wsB.cell(r, c).value = None
for r in range(7, 409):
    sid = wsB.cell(r, 1).value
    if sid and sid in ta:
        si = sd[sid]["sup_idx"]
        sq = max(1, int(round(sd[sid]["order"] * mr[si])))
        ti = ta[sid]
        for w in range(24): wsB.cell(r, 2 + w * 8 + ti).value = sq
wbB.save("附件B 转运方案数据结果.xlsx")
print("附件B Q4 OK")

# 验证Q2/Q3未被修改
for sn in wbA.sheetnames:
    ws = wbA[sn]
    cnt = sum(1 for r in range(7, 409) if ws.cell(r, 2).value)
    if cnt > 0: print("  {}: {}家".format(sn, cnt))
for sn in wbB.sheetnames:
    ws = wbB[sn]
    cnt = sum(1 for r in range(7, 409) for t in range(8) if ws.cell(r, 2 + t).value)
    if cnt > 0: print("  {}: {}家".format(sn, cnt))
print("完成。")
